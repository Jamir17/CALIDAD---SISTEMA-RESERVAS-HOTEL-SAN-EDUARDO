from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, Response
from bd import obtener_conexion
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill

reservas_cliente_bp = Blueprint('reservas_cliente', __name__, url_prefix='/historial-cliente')

def requiere_login_rol(roles_permitidos):
    """Decorador para proteger rutas por rol de usuario."""
    def wrapper(fn):
        def inner(*args, **kwargs):
            if 'usuario_id' not in session:
                flash("Primero inicia sesión.", "error")
                return redirect(url_for("usuarios.iniciosesion"))
            if session.get("rol") not in roles_permitidos:
                flash("No tienes permiso para acceder a esta sección.", "error")
                return redirect(url_for("admin.dashboard"))
            return fn(*args, **kwargs)
        inner.__name__ = fn.__name__
        return inner
    return wrapper

@reservas_cliente_bp.route('/')
@requiere_login_rol({1, 2}) # Solo Admin y Recepcionista
def historial_reservas_cliente():
    """Renderiza la página principal del historial de reservas."""
    return render_template('reservas_cliente.html')

@reservas_cliente_bp.route('/api/buscar', methods=['POST'])
@requiere_login_rol({1, 2})
def buscar_historial_cliente():
    """
    Busca el historial de reservas de un cliente según los filtros proporcionados.
    Devuelve los datos en formato JSON para ser consumidos por JavaScript.
    """
    data = request.get_json()
    search_term = data.get('searchTerm', '').strip()
    estado = data.get('estado')
    fecha_desde = data.get('fechaDesde')
    fecha_hasta = data.get('fechaHasta')

    if not search_term:
        return jsonify({
            "ok": False,
            "message": "Por favor, ingrese un nombre o documento para buscar."
        }), 400

    conn = obtener_conexion()
    try:
        with conn.cursor() as cursor:
            # 1. Buscar el cliente
            cursor.execute("""
                SELECT id_cliente, nombres, apellidos, num_documento
                FROM clientes
                WHERE CONCAT(nombres, ' ', apellidos) LIKE %s OR num_documento LIKE %s
                LIMIT 1
            """, (f"%{search_term}%", f"%{search_term}%"))
            cliente = cursor.fetchone()

            if not cliente:
                return jsonify({"ok": False, "message": "No se encontró ningún cliente con esos datos."})

            # 2. Construir la consulta para las reservas del cliente encontrado
            query = """
                SELECT
                    r.id_reserva,
                    r.codigo_confirmacion,
                    h.numero AS habitacion_numero,
                    r.fecha_entrada,
                    r.fecha_salida,
                    r.noches,
                    r.estado,
                    COALESCE(r.total, f.total, 0) AS monto
                FROM reservas r
                JOIN habitaciones h ON r.id_habitacion = h.id_habitacion
                LEFT JOIN facturacion f ON r.id_reserva = f.id_reserva
                WHERE r.id_cliente = %s
            """
            params = [cliente['id_cliente']]

            if estado and estado != "Todos":
                query += " AND r.estado = %s"
                params.append(estado)

            if fecha_desde:
                query += " AND r.fecha_entrada >= %s"
                params.append(fecha_desde)

            if fecha_hasta:
                query += " AND r.fecha_salida <= %s"
                params.append(fecha_hasta)

            query += " ORDER BY r.fecha_entrada DESC"

            cursor.execute(query, params)
            reservas = cursor.fetchall()

            # Formatear fechas para JSON
            for r in reservas:
                r['fecha_entrada'] = r['fecha_entrada'].strftime('%d/%m/%Y') if r['fecha_entrada'] else '-'
                r['fecha_salida'] = r['fecha_salida'].strftime('%d/%m/%Y') if r['fecha_salida'] else '-'

            return jsonify({
                "ok": True,
                "cliente": cliente,
                "reservas": reservas
            })

    except Exception as e:
        print(f"Error en buscar_historial_cliente: {e}")
        return jsonify({"ok": False, "message": "Ocurrió un error en el servidor."}), 500
    finally:
        if conn:
            conn.close()

@reservas_cliente_bp.route('/api/exportar-excel', methods=['POST'])
@requiere_login_rol({1, 2})
def exportar_excel_historial():
    """
    Exporta el historial de reservas de un cliente a un archivo Excel (.xlsx)
    con los filtros y estilos solicitados.
    """
    data = request.get_json()
    search_term = data.get('searchTerm', '').strip()
    estado = data.get('estado')
    fecha_desde = data.get('fechaDesde')
    fecha_hasta = data.get('fechaHasta')

    if not search_term:
        return jsonify({"ok": False, "message": "Se requiere un término de búsqueda."}), 400

    conn = obtener_conexion()
    try:
        with conn.cursor() as cursor:
            # 1. Buscar cliente (similar a la búsqueda normal)
            cursor.execute("SELECT id_cliente, nombres, apellidos FROM clientes WHERE CONCAT(nombres, ' ', apellidos) LIKE %s OR num_documento LIKE %s LIMIT 1", (f"%{search_term}%", f"%{search_term}%"))
            cliente = cursor.fetchone()
            if not cliente:
                return jsonify({"ok": False, "message": "Cliente no encontrado."}), 404

            # 2. Obtener las reservas con los mismos filtros
            query = """
                SELECT r.codigo_confirmacion, h.numero, r.fecha_entrada, r.fecha_salida, r.noches, r.estado, COALESCE(r.total, f.total, 0) AS monto
                FROM reservas r
                JOIN habitaciones h ON r.id_habitacion = h.id_habitacion
                LEFT JOIN facturacion f ON r.id_reserva = f.id_reserva
                WHERE r.id_cliente = %s
            """
            params = [cliente['id_cliente']]
            if estado and estado != "Todos":
                query += " AND r.estado = %s"
                params.append(estado)
            if fecha_desde:
                query += " AND r.fecha_entrada >= %s"
                params.append(fecha_desde)
            if fecha_hasta:
                query += " AND r.fecha_salida <= %s"
                params.append(fecha_hasta)
            query += " ORDER BY r.fecha_entrada DESC"
            cursor.execute(query, params)
            reservas = cursor.fetchall()

        # 3. Crear el archivo Excel en memoria
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HistorialReservas"

        # 4. Definir estilos para la cabecera
        header_font = Font(name='Calibri', bold=True, color='000000')
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Amarillo

        # 5. Escribir cabeceras y aplicar estilos
        headers = ["Codigo", "Habitacion", "Entrada", "Salida", "Noches", "Estado", "Monto"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill

        # 6. Escribir datos de las reservas
        for row_num, r in enumerate(reservas, 2):
            ws.cell(row=row_num, column=1, value=r['codigo_confirmacion'] or 'N/A')
            ws.cell(row=row_num, column=2, value=f"Hab. {r['numero']}")
            ws.cell(row=row_num, column=3, value=r['fecha_entrada'].strftime('%d/%m/%Y') if r['fecha_entrada'] else '-')
            ws.cell(row=row_num, column=4, value=r['fecha_salida'].strftime('%d/%m/%Y') if r['fecha_salida'] else '-')
            ws.cell(row=row_num, column=5, value=r['noches'])
            ws.cell(row=row_num, column=6, value=r['estado'])
            ws.cell(row=row_num, column=7, value=float(r['monto'] or 0))

        # 7. Guardar el libro en un stream de bytes y devolverlo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        cliente_nombre_archivo = f"{cliente['nombres']}_{cliente['apellidos']}".replace(" ", "_")
        return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment;filename=Historial_{cliente_nombre_archivo}.xlsx'})

    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return jsonify({"ok": False, "message": "Error interno al generar el archivo Excel."}), 500
    finally:
        if conn:
            conn.close()